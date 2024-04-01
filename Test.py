import webbrowser
from et_xmlfile import xmlfile
from openpyxl import load_workbook
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
import xlsxwriter
from kivy.uix.image import Image
from kivy.metrics import inch
class DeltaPopLocaties(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.cols = 1
        self.inside = BoxLayout()
        self.inside.cols = 3
        self.orientation = "vertical"

        def on_size(self, instance, value):
            # run parent on_size manually
            super(TextInput, self).on_size(instance, value)
            self.font_size = self.height * 0.8

        self.label_search = Label(text='Type de Naam van jou gewenste pop \n Bijvoorbeeld AB-02\n Hou de zoek knopje ingedrukt:')
        self.add_widget(self.label_search)

        self.entry_search = TextInput(multiline=False)
        self.inside.add_widget(self.entry_search)

        self.button_search = Button(text="Zoeken")
        self.button_search.bind(on_press=self.search_excel)
        self.inside.add_widget(self.button_search)
        self.add_widget(self.inside)
    def search_excel(self, instance):
        search_query = self.entry_search.text
        if not search_query:
            popup = Popup(title='Error', content=Label(text='A.U.B Type de naam van de pop.', multiline=False), size_hint=(None, None),
                          size=(400, 400))
            popup.open()
            return

        try:
            wb = load_workbook("data.xlsx")
            sheet = wb.active
            result = None

            for row in sheet.iter_rows(values_only=True):
                if search_query.lower() in str(row[0]).lower():
                    result = row
                    break

            if result:
                result_dict = {}
                for i in range(1, len(result)):
                    title = sheet.cell(row=1, column=i + 1).value
                    if title in ["Adres", "Aanvulling Adres",]:
                        result_dict[title] = result[i]

                result_str = "\n".join([f"{key}: {value}" for key, value in result_dict.items()])
                popup = Popup(title='Search Result', content=ScrollView(size_hint=(2,None), size=(600, 600)),
                              size_hint=(1, None), size=(800, 400))
                popup.content.add_widget(Label(text=result_str))
                popup.open()

                # Add hyperlink button
                if 'Adres' in result_dict:
                    address = result_dict['Adres']
                    hyperlink_button = Button(text=f"Open Google Maps",size_hint=(None, None),
                                              size=(800, 700))
                    hyperlink_button.bind(on_press=lambda instance: self.open_google_maps(address))
                    # Wrap the button in a BoxLayout to center it horizontally
                    button_layout = BoxLayout(orientation='horizontal',size_hint=(None, None), size=(400, 100))
                    button_layout.add_widget(Label())  # Empty label to push the button to the center
                    button_layout.add_widget(hyperlink_button)
                    self.add_widget(button_layout)

            else:
                popup = Popup(title='Search Result', content=Label(text='Search query not found.'),
                              size_hint=(None, None), size=(700, 400))
                popup.open()
        except FileNotFoundError:
            popup = Popup(title='Error', content=Label(text='Excel file not found.'), size_hint=(None, None),
                          size=(600, 200))
            popup.open()

    def open_google_maps(self, address):
        if address:
            hyperlink = f"https://www.google.com/maps/search/{address}"
            webbrowser.open(hyperlink)
            return Image(source='log.jpg')
class ExcelLocationSearchApp(App):
    def build(self):
        return DeltaPopLocaties()

if __name__ == "__main__":
    ExcelLocationSearchApp().run()
                  self.add_widget(button_layout)

            else:
                popup = Popup(title='Search Result', content=Label(text='Search query not found.'),
                              size_hint=(None, None), size=(600, 200))
                popup.open()
        except FileNotFoundError:
            popup = Popup(title='Error', content=Label(text='Excel file not found.'), size_hint=(None, None),
                          size=(400, 200))
            popup.open()

    def open_google_maps(self, address):
        if address:
            hyperlink = f"https://www.google.com/maps/search/{address}"
            webbrowser.open(hyperlink)


class ExcelLocationSearchApp(App):
    def build(self):
        return DeltaPopLocaties()


if __name__ == "__main__":
    ExcelLocationSearchApp().run()
